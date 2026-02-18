import os
import sys
import re
import tempfile
import pandas as pd
from copy import copy
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# =============================================================================
# ███████╗████████╗███████╗██████╗      ██╗
# ██╔════╝╚══██╔══╝██╔════╝██╔══██╗    ███║
# ███████╗   ██║   █████╗  ██████╔╝    ╚██║
# ╚════██║   ██║   ██╔══╝  ██╔═══╝      ██║
# ███████║   ██║   ███████╗██║          ██║
# ╚══════╝   ╚═╝   ╚══════╝╚═╝          ╚═╝
#
# STEP 1 — CLEAN & CONVERT  (formerly Code 1)
# Reads the raw survey output xlsx and produces a clean Question_Options file.
# =============================================================================

QUESTION_RE = re.compile(r'^\s*Q\s*(\d+)\s*[\.\)\:\-]?\s*(.*)$', re.IGNORECASE)

METADATA_TOKENS = {
    "answer choices", "answered", "skipped", "responses", "response",
    "average", "total", "base", "count", "%", "weighted base",
    "unweighted base"
}


def clean_value(v):
    if pd.isna(v) or v == '':
        return ""
    try:
        f = float(v)
        if f.is_integer():
            return str(int(f))
        return str(f)
    except Exception:
        return str(v).strip()


def is_metadata_text(s):
    if not s:
        return False
    return s.strip().lower() in METADATA_TOKENS


def is_question_row(text):
    return bool(QUESTION_RE.match(text or ""))


def parse_question_row(text):
    m = QUESTION_RE.match(text or "")
    if not m:
        return None, None
    return f"Q{int(m.group(1))}", m.group(2).strip()


# ------------------------------------------------------------------
# BIPOLAR SCALE DETECTION
# ------------------------------------------------------------------

def is_bipolar_sub_label(first_cell, next_row):
    if not first_cell or is_question_row(first_cell):
        return False
    if " or " not in first_cell.lower():
        return False

    next_col0 = str(next_row.iloc[0]).strip() if pd.notna(next_row.iloc[0]) else ""
    next_col1 = str(next_row.iloc[1]).strip() if len(next_row) > 1 and pd.notna(next_row.iloc[1]) else ""

    if next_col0 != "":
        return False
    if next_col1 == "":
        return False

    try:
        float(next_col1)
        return False
    except ValueError:
        pass

    return True


def detect_bipolar_question(df, q_start_idx):
    sub_labels = []
    pole_1 = None
    pole_2 = None

    for idx in range(q_start_idx + 1, len(df)):
        first = str(df.iloc[idx, 0]).strip() if pd.notna(df.iloc[idx, 0]) else ""

        if is_question_row(first):
            break

        if idx + 1 < len(df) and is_bipolar_sub_label(first, df.iloc[idx + 1]):
            sub_labels.append(first)

            if pole_1 is None:
                poles_row = df.iloc[idx + 1]
                poles = []
                for c in range(1, min(20, len(poles_row))):
                    val = poles_row.iloc[c]
                    if pd.notna(val):
                        val_str = str(val).strip()
                        if val_str.lower() == "total":
                            break
                        if not is_metadata_text(val_str):
                            try:
                                float(val_str)
                            except ValueError:
                                poles.append(val_str)
                if len(poles) >= 2:
                    pole_1, pole_2 = poles[0], poles[1]

    return sub_labels, pole_1, pole_2


# ------------------------------------------------------------------
# TYPE DETECTION
# ------------------------------------------------------------------

def get_auto_type(q_text, rank_labels, options, is_bipolar=False):
    if is_bipolar:
        return "Bipolar"

    if rank_labels:
        return "Matrix"

    q_lower = q_text.lower()

    SINGLE_KEYWORDS = [
        "single selection", "single choice", "single select",
        "select one", "[single", "unaided single"
    ]
    MULTIPLE_KEYWORDS = [
        "multiple selection", "multiple choice", "multiple select",
        "multi selection", "multi select", "[multiple", "[multi",
        "aided multiple", "unaided multiple"
    ]

    for keyword in SINGLE_KEYWORDS:
        if keyword in q_lower:
            return "Single"
    for keyword in MULTIPLE_KEYWORDS:
        if keyword in q_lower:
            return "Multiple"

    return ""


# ------------------------------------------------------------------
# NPS AUTO EXPANSION
# ------------------------------------------------------------------

def is_scale_value(val_clean):
    if val_clean.isdigit():
        return True
    if "-" in val_clean:
        return True
    try:
        f = float(val_clean)
        if f == int(f):
            return True
    except ValueError:
        pass
    return False


def expand_nps_if_needed(q_text, options):
    q_lower = q_text.lower()

    if "1 to 10" in q_lower or "1-10" in q_lower or "scale of 0 to 10" in q_lower:
        labels = {}

        match_1 = re.search(r'1\s+means?\s+["\u201c]?([^"\u201d]+)', q_text, re.IGNORECASE)
        if match_1:
            labels["1"] = f"1- {match_1.group(1).strip()}"

        match_10 = re.search(r'10\s+means?\s+["\u201c]?([^"\u201d]+)', q_text, re.IGNORECASE)
        if match_10:
            labels["10"] = f"10- {match_10.group(1).strip()}"

        new_options = []
        for i in range(1, 11):
            key = str(i)
            if key in labels:
                new_options.append(labels[key])
            else:
                new_options.append(key)

        return new_options

    return options


# ------------------------------------------------------------------
# HORIZONTAL SCALE DETECTION
# ------------------------------------------------------------------

def detect_horizontal_scale(df, idx):
    def extract_scale(row):
        scale_values = []
        for col in range(1, min(30, len(row))):
            val = row.iloc[col]
            if pd.notna(val):
                val_clean = str(val).strip()
                if val_clean.lower() in ("total", "weighted average"):
                    break
                if is_scale_value(val_clean):
                    scale_values.append(val_clean)
        return scale_values

    scale = extract_scale(df.iloc[idx])
    if len(scale) >= 3:
        return scale

    if idx + 1 < len(df):
        scale = extract_scale(df.iloc[idx + 1])
        if len(scale) >= 3:
            return scale

    return []


# ------------------------------------------------------------------
# MAIN CLEAN + CONVERT  (Step 1 entry point)
# ------------------------------------------------------------------

def clean_and_convert(input_path, output_path):
    """
    Reads the raw survey Excel file (input_path) and writes a clean
    Question_Options xlsx to output_path.  This is Step 1.
    """
    print(f"\n📂 Reading raw data from: {input_path}")
    df = pd.read_excel(input_path, header=None)

    questions = []
    current_q_text = None
    current_options = []
    current_rank_labels = []
    current_is_bipolar = False
    in_question = False

    idx = 0
    while idx < len(df):

        first_cell = str(df.iloc[idx, 0]).strip() if pd.notna(df.iloc[idx, 0]) else ""

        # ── Empty row ──────────────────────────────────────────────────────
        if first_cell == "":
            if current_q_text and (current_options or current_rank_labels):
                if not current_rank_labels:
                    current_options = expand_nps_if_needed(current_q_text, current_options)
                questions.append({
                    "q_text": current_q_text,
                    "options": current_options.copy(),
                    "rank_labels": current_rank_labels.copy(),
                    "is_bipolar": current_is_bipolar,
                })

            current_q_text = None
            current_options = []
            current_rank_labels = []
            current_is_bipolar = False
            in_question = False
            idx += 1
            continue

        # ── Question header ────────────────────────────────────────────────
        if is_question_row(first_cell):

            if current_q_text and (current_options or current_rank_labels):
                if not current_rank_labels:
                    current_options = expand_nps_if_needed(current_q_text, current_options)
                questions.append({
                    "q_text": current_q_text,
                    "options": current_options.copy(),
                    "rank_labels": current_rank_labels.copy(),
                    "is_bipolar": current_is_bipolar,
                })

            _, q_text = parse_question_row(first_cell)
            current_q_text = q_text
            current_options = []
            current_rank_labels = []
            current_is_bipolar = False
            in_question = True

            # ── BIPOLAR ────────────────────────────────────────────────────
            bipolar_labels, pole_1, pole_2 = detect_bipolar_question(df, idx)
            if bipolar_labels:
                current_options = bipolar_labels
                current_rank_labels = [pole_1 or "Pole_1", pole_2 or "Pole_2"]
                current_is_bipolar = True

                skip_to = idx + 1
                while skip_to < len(df):
                    ahead_cell = str(df.iloc[skip_to, 0]).strip() if pd.notna(df.iloc[skip_to, 0]) else ""
                    if is_question_row(ahead_cell):
                        break
                    skip_to += 1
                idx = skip_to
                continue

            # ── MATRIX detection ───────────────────────────────────────────
            if idx + 1 < len(df):
                next_row = df.iloc[idx + 1]
                ranks = []

                for col in range(1, min(30, len(next_row))):
                    val = next_row.iloc[col]
                    if pd.notna(val):
                        val_clean = str(val).strip()
                        if val_clean.lower() == "total":
                            break
                        if not is_metadata_text(val_clean):
                            ranks.append(val_clean)

                if len(ranks) >= 2:
                    attributes = []
                    for r in range(idx + 2, len(df)):
                        val = df.iloc[r, 0]
                        if pd.isna(val):
                            break
                        attr = str(val).strip()
                        if is_question_row(attr):
                            break
                        if not is_metadata_text(attr):
                            attributes.append(attr)

                    if len(attributes) >= 2:
                        current_options = attributes
                        current_rank_labels = ranks

            # ── Horizontal scale ───────────────────────────────────────────
            if not current_rank_labels:
                scale = detect_horizontal_scale(df, idx)
                if scale:
                    current_options = scale

            idx += 1
            continue

        # ── Inside a question (standard option rows) ───────────────────────
        elif in_question and not current_rank_labels:
            if not is_metadata_text(first_cell) and first_cell.lower() != "answer choices":
                current_options.append(first_cell)

        idx += 1

    # Save last question
    if current_q_text and (current_options or current_rank_labels):
        if not current_rank_labels:
            current_options = expand_nps_if_needed(current_q_text, current_options)
        questions.append({
            "q_text": current_q_text,
            "options": current_options.copy(),
            "rank_labels": current_rank_labels.copy(),
            "is_bipolar": current_is_bipolar,
        })

    # ── TYPE VALIDATION ────────────────────────────────────────────────────
    missing_type_questions = []

    for q in questions:
        detected_type = get_auto_type(q["q_text"], q["rank_labels"], q["options"], q.get("is_bipolar", False))
        if detected_type == "":
            missing_type_questions.append(q)

    bipolar_count = sum(1 for q in questions if q.get("is_bipolar", False))
    matrix_count  = sum(1 for q in questions if q["rank_labels"] and not q.get("is_bipolar", False))

    print("\n📊 Type Detection Summary:")
    print(f"   Total Questions : {len(questions)}")
    print(f"   Bipolar         : {bipolar_count}")
    print(f"   Matrix          : {matrix_count}")
    print(f"   Auto-detected   : {len(questions) - len(missing_type_questions)}")
    print(f"   Needs manual    : {len(missing_type_questions)}")

    if missing_type_questions:
        print("\n⚠️  MANUAL TYPE ENTRY REQUIRED")
        for q in missing_type_questions:
            print(f"\nQuestion: {q['q_text']}")
            print(f"Options:  {', '.join(q['options'][:3])}...")
            while True:
                user_input = input("Enter type (Single/Multiple): ").strip().capitalize()
                if user_input in ["Single", "Multiple"]:
                    q["manual_type"] = user_input
                    break
                else:
                    print("❌ Invalid input. Please enter Single or Multiple.")

    # ── WRITE OPTIONS FILE ─────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Question_Options"

    max_ranks = max((len(q["rank_labels"]) for q in questions), default=0)

    headers = ["Question Text", "Option", "Type"]
    for i in range(1, max_ranks + 1):
        headers.append(f"Rank_{i}")

    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF", name="Arial")
        cell.fill = PatternFill(start_color="366092", fill_type="solid")

    row = 2
    for q in questions:
        q_type = get_auto_type(q["q_text"], q["rank_labels"], q["options"], q.get("is_bipolar", False))
        if q_type == "":
            q_type = q.get("manual_type", "")

        for opt in q["options"]:
            ws.cell(row, 1, q["q_text"]).font = Font(name="Arial")
            ws.cell(row, 2, clean_value(opt)).font = Font(name="Arial")
            ws.cell(row, 3, q_type).font = Font(name="Arial")

            for r_idx, rank in enumerate(q["rank_labels"]):
                ws.cell(row, 4 + r_idx, clean_value(rank)).font = Font(name="Arial")

            row += 1

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

    wb.save(output_path)
    print(f"\n✅ Step 1 complete — Options file saved: {output_path}")


# =============================================================================
# ███████╗████████╗███████╗██████╗     ██████╗
# ██╔════╝╚══██╔══╝██╔════╝██╔══██╗    ╚════██╗
# ███████╗   ██║   █████╗  ██████╔╝     █████╔╝
# ╚════██║   ██║   ██╔══╝  ██╔═══╝     ██╔═══╝
# ███████║   ██║   ███████╗██║         ███████╗
# ╚══════╝   ╚═╝   ╚══════╝╚═╝         ╚══════╝
#
# STEP 2 — SURVEY DATABOOK GENERATOR  (formerly Code 2)
# Uses the options file produced in Step 1 to build the full databook.
# =============================================================================

class SurveyDatabookV2:
    """
    Survey Databook Generator V3
    Supports: Single, Multiple, Matrix, Bipolar question types.
    """

    def __init__(self, data_file_path, options_file_path):
        self.data_file_path    = data_file_path
        self.options_file_path = options_file_path
        self.df                = None
        self.df_full           = None
        self.metadata_columns  = []
        self.question_map      = {}
        self.question_config   = {}

    # ------------------------------------------------------------------
    # HELPERS
    # ------------------------------------------------------------------

    def get_excel_column_letter(self, col_num):
        letter = ''
        while col_num > 0:
            col_num, remainder = divmod(col_num - 1, 26)
            letter = chr(65 + remainder) + letter
        return letter

    # ------------------------------------------------------------------
    # LOAD OPTIONS FILE
    # ------------------------------------------------------------------

    def load_options_file(self):
        if not self.options_file_path:
            print("❌ Options file is REQUIRED")
            sys.exit(1)

        df_opts = pd.read_excel(self.options_file_path)

        required_cols = ['Question Text', 'Option', 'Type']
        missing = [c for c in required_cols if c not in df_opts.columns]
        if missing:
            print(f"❌ Options file missing columns: {missing}")
            sys.exit(1)

        cols     = list(df_opts.columns)
        type_idx = cols.index('Type')

        for question_text, group in df_opts.groupby('Question Text', sort=False):
            question_text = str(question_text).strip()
            group         = group.drop_duplicates(subset=['Option'], keep='first')

            q_types = group['Type'].dropna().unique()
            q_type  = str(q_types[0]).strip() if len(q_types) > 0 else 'Single'

            options = [str(opt).strip() for opt in group['Option'].dropna().tolist()]

            # ── BIPOLAR ──────────────────────────────────────────────────────
            if q_type.lower() == 'bipolar':
                pole_pairs = []
                for opt in options:
                    if ' or ' in opt.lower():
                        parts = re.split(r'\s+or\s+', opt, maxsplit=1, flags=re.IGNORECASE)
                        pole_pairs.append((parts[0].strip(), parts[1].strip()))
                    else:
                        pole_pairs.append((opt, ''))

                self.question_config[question_text] = {
                    'type':        'bipolar',
                    'options':     options,
                    'pole_pairs':  pole_pairs,
                    'rank_labels': []
                }
                print(f"   🔵 Bipolar : {question_text[:60]}... → {len(options)} dimensions")
                continue

            # ── MATRIX ───────────────────────────────────────────────────────
            rank_labels = []
            if 'matrix' in q_type.lower():
                extra_cols = cols[type_idx + 1:]
                first_row  = group.iloc[0]
                for col_name in extra_cols:
                    val = first_row[col_name]
                    if pd.isna(val):
                        continue
                    label = str(val).strip()
                    if not label:
                        continue
                    if label.endswith('.0'):
                        label = label[:-2]
                    rank_labels.append(label)
                print(
                    f"   📋 Matrix  : {question_text[:50]}... → "
                    f"{len(options)} rows, {len(rank_labels)} cols"
                )

            self.question_config[question_text] = {
                'type':        q_type.lower(),
                'options':     options,
                'rank_labels': rank_labels
            }

        print(f"\n✅ Loaded config for {len(self.question_config)} questions")
        type_counts = {}
        for config in self.question_config.values():
            t = config['type']
            type_counts[t] = type_counts.get(t, 0) + 1
        print("   Question types:")
        for t, c in sorted(type_counts.items()):
            print(f"     - {t}: {c}")

    # ------------------------------------------------------------------
    # LOAD DATA
    # ------------------------------------------------------------------

    def load_data(self):
        if self.data_file_path.endswith('.csv'):
            df_full = pd.read_csv(self.data_file_path, header=None)
        else:
            df_full = pd.read_excel(self.data_file_path, header=None)

        self.df_full = df_full

        if self.data_file_path.endswith('.csv'):
            df_with_headers = pd.read_csv(self.data_file_path)
        else:
            df_with_headers = pd.read_excel(self.data_file_path)

        if 'Respondent ID' in df_with_headers.columns:
            self.df = df_with_headers[df_with_headers['Respondent ID'].notna()].reset_index(drop=True)
            print(f"✅ Loaded {len(self.df)} valid responses")
        else:
            self.df = df_with_headers
            print(f"✅ Loaded {len(self.df)} responses")

    # ------------------------------------------------------------------
    # MAP QUESTIONS
    # ------------------------------------------------------------------

    def identify_structure(self):
        exact_metadata = [
            'respondent id', 'collector id', 'start date', 'end date',
            'ip address', 'email address', 'first name', 'last name',
            'custom data 1', 'custom data 2', 'custom data'
        ]
        exact_exclude = [
            'status', 'test/garbage entry', 'reason for rejection',
            'contact no.', 'contact no', 'surveyor name', 'tiers'
        ]

        current_question  = None
        question_counter  = 0

        for idx, col in enumerate(self.df.columns):
            col_lower = str(col).lower().strip()

            if col_lower in exact_metadata or col_lower in exact_exclude:
                self.metadata_columns.append(col)
                continue

            if 'unnamed' in col_lower:
                if current_question is not None:
                    self.question_map[current_question]['unnamed_cols'].append(idx)
            else:
                question_counter  += 1
                current_question   = question_counter
                self.question_map[current_question] = {
                    'question_text': col,
                    'main_col_idx':  idx,
                    'unnamed_cols':  []
                }

        print(f"✅ Identified {len(self.question_map)} questions")

    # ------------------------------------------------------------------
    # BASIC PROCESSORS
    # ------------------------------------------------------------------

    def merge_unnamed_columns(self, question_info):
        main_col_idx        = question_info['main_col_idx']
        unnamed_col_indices = question_info['unnamed_cols']

        merged_data = []
        for row_idx in range(len(self.df)):
            values   = []
            main_val = self.df.iloc[row_idx, main_col_idx]
            if pd.notna(main_val) and str(main_val).strip():
                values.append(str(main_val).strip())
            for unnamed_idx in unnamed_col_indices:
                unnamed_val = self.df.iloc[row_idx, unnamed_idx]
                if pd.notna(unnamed_val) and str(unnamed_val).strip():
                    values.append(str(unnamed_val).strip())
            merged_data.append(', '.join(values) if values else None)

        return pd.Series(merged_data)

    def split_multi_select_options(self, text):
        pattern = r',\s*(?![^()]*\))'
        parts   = re.split(pattern, str(text))
        return [p.strip() for p in parts if p.strip()]

    def process_single_select(self, series, options):
        series_clean = series.dropna()
        series_clean = series_clean[series_clean != 'Response']
        series_clean = series_clean[series_clean != 'Open-Ended Response']
        series_clean = series_clean[series_clean.astype(str).str.strip() != '']

        option_counts = series_clean.value_counts()
        total_base    = len(series_clean)

        results = []
        if options:
            for option in options:
                count = option_counts.get(option, 0)
                results.append({
                    'option':     option,
                    'count':      int(count),
                    'percentage': round((count / total_base) * 100, 1) if total_base > 0 else 0.0,
                    'base':       total_base
                })
        else:
            for option, count in option_counts.items():
                results.append({
                    'option':     str(option),
                    'count':      int(count),
                    'percentage': round((count / total_base) * 100, 1) if total_base > 0 else 0.0,
                    'base':       total_base
                })
            results.sort(key=lambda x: x['count'], reverse=True)

        return results

    def process_multi_select(self, series, options):
        all_options = []
        for val in series.dropna():
            val_str = str(val).strip()
            if val_str and val_str not in ['Response', 'Open-Ended Response']:
                opts = self.split_multi_select_options(val_str)
                all_options.extend(opts)

        option_counts = pd.Series(all_options).value_counts() if all_options else pd.Series()
        total_base    = len(series.dropna())

        results = []
        if options:
            for option in options:
                count = option_counts.get(option, 0)
                results.append({
                    'option':     option,
                    'count':      int(count),
                    'percentage': round((count / total_base) * 100, 1) if total_base > 0 else 0.0,
                    'base':       total_base
                })
        else:
            for option, count in option_counts.items():
                results.append({
                    'option':     option,
                    'count':      int(count),
                    'percentage': round((count / total_base) * 100, 1) if total_base > 0 else 0.0,
                    'base':       total_base
                })
            results.sort(key=lambda x: x['count'], reverse=True)

        return results

    # ------------------------------------------------------------------
    # BIPOLAR PROCESSOR
    # ------------------------------------------------------------------

    def process_bipolar(self, question_info, options, pole_pairs):
        main_col_idx = question_info['main_col_idx']
        unnamed_cols = question_info['unnamed_cols']
        all_col_idxs = [main_col_idx] + unnamed_cols

        total_responses = len(self.df)

        dim_to_col = {}
        for col_idx in all_col_idxs:
            if col_idx >= len(self.df_full.columns):
                continue
            sub_hdr = self.df_full.iloc[1, col_idx]
            if pd.isna(sub_hdr):
                continue
            sub_hdr_str = str(sub_hdr).strip()

            if ' - ' in sub_hdr_str:
                dimension = sub_hdr_str.split(' - ', 1)[1].strip()
            else:
                dimension = sub_hdr_str

            dim_to_col[dimension] = col_idx

        print(f"      Bipolar dim→col mapping: {dim_to_col}")

        bipolar_rows = []

        for opt, (pole_1, pole_2) in zip(options, pole_pairs):
            col_idx = dim_to_col.get(opt)

            if col_idx is None:
                for key, val in dim_to_col.items():
                    if opt.lower() in key.lower() or key.lower() in opt.lower():
                        col_idx = val
                        break

            if col_idx is None:
                print(f"      ⚠ No raw column found for dimension: '{opt}'")
                bipolar_rows.append({
                    'option':   opt,
                    'pole_1':   pole_1,
                    'pole_2':   pole_2,
                    'col_idx':  None,
                    'count_p1': 0,
                    'count_p2': 0,
                    'n':        0,
                })
                continue

            col_data = self.df.iloc[:, col_idx].dropna()
            col_data = col_data[col_data.astype(str).str.strip() != '']

            count_p1 = int((col_data == pole_1).sum())
            count_p2 = int((col_data == pole_2).sum())
            n        = len(col_data)

            bipolar_rows.append({
                'option':   opt,
                'pole_1':   pole_1,
                'pole_2':   pole_2,
                'col_idx':  col_idx,
                'count_p1': count_p1,
                'count_p2': count_p2,
                'n':        n,
            })

        return {
            'is_bipolar': True,
            'rows':       bipolar_rows,
            'base':       total_responses,
        }

    # ------------------------------------------------------------------
    # MATRIX PROCESSOR
    # ------------------------------------------------------------------

    def process_matrix(self, question_info, options, rank_labels):
        main_col_idx        = question_info['main_col_idx']
        unnamed_col_indices = question_info['unnamed_cols']
        question_text       = question_info['question_text']

        print(f"\n   🔍 Matrix: {question_text[:60]}...")
        print(f"      {len(options)} rows × {len(rank_labels)} cols")

        all_cols = [main_col_idx] + unnamed_col_indices

        combined_separators = 0
        sample_cols         = all_cols[:min(10, len(all_cols))]
        for col_idx in sample_cols:
            header_val = self.df_full.iloc[1, col_idx]
            if pd.notna(header_val) and ' - ' in str(header_val):
                combined_separators += 1

        is_multi_select_matrix = combined_separators >= max(1, len(sample_cols) * 0.5)

        column_mapping = []

        if is_multi_select_matrix:
            print("      🔧 MATRIX MULTI-SELECT")
            horizontal_set = set(rank_labels)
            for col_idx in all_cols:
                header_val = self.df_full.iloc[1, col_idx]
                if pd.isna(header_val):
                    continue
                raw = str(header_val).strip()
                if ' - ' not in raw:
                    continue
                left, right = [p.strip() for p in raw.split(' - ', 1)]
                if left in horizontal_set:
                    rank_label, attribute = left, right
                elif right in horizontal_set:
                    rank_label, attribute = right, left
                else:
                    rank_label, attribute = left, right
                column_mapping.append({'col_idx': col_idx, 'attribute': attribute, 'rank_label': rank_label})
        else:
            print("      🔧 SIMPLE MATRIX SINGLE-SELECT")
            for col_idx in all_cols:
                header_val = self.df_full.iloc[1, col_idx]
                if pd.isna(header_val):
                    continue
                attribute = str(header_val).strip()
                if not attribute:
                    continue
                column_mapping.append({'col_idx': col_idx, 'attribute': attribute, 'rank_label': None})

        total_responses = len(self.df)
        matrix_data     = []

        for option_attribute in options:
            row_data = {'attribute': option_attribute}
            for rank_label in rank_labels:
                count = 0
                if is_multi_select_matrix:
                    matching_cols = [
                        cm['col_idx'] for cm in column_mapping
                        if cm['attribute'] == option_attribute and cm['rank_label'] == rank_label
                    ]
                    for row_idx in range(len(self.df)):
                        for col_idx in matching_cols:
                            val = self.df.iloc[row_idx, col_idx]
                            if pd.notna(val) and str(val).strip():
                                count += 1
                else:
                    matching_cols = [
                        cm['col_idx'] for cm in column_mapping
                        if cm['attribute'] == option_attribute
                    ]
                    for row_idx in range(len(self.df)):
                        for col_idx in matching_cols:
                            val = self.df.iloc[row_idx, col_idx]
                            if pd.notna(val) and str(val).strip() == str(rank_label):
                                count += 1

                row_data[rank_label] = {
                    'count':      int(count),
                    'percentage': round((count / total_responses) * 100, 1) if total_responses > 0 else 0.0
                }
            matrix_data.append(row_data)

        return {
            'is_matrix':              True,
            'attributes':             options,
            'rank_labels':            rank_labels,
            'data':                   matrix_data,
            'base':                   total_responses,
            'column_mapping':         column_mapping,
            'is_multi_select_matrix': is_multi_select_matrix
        }

    # ------------------------------------------------------------------
    # ANALYZE ALL QUESTIONS
    # ------------------------------------------------------------------

    def analyze_survey(self):
        analysis = {
            'total_responses': len(self.df),
            'total_questions': len(self.question_map),
            'questions':       []
        }

        for q_num, q_info in self.question_map.items():
            question_text = q_info['question_text']
            config        = self.question_config.get(question_text)

            if not config:
                print(f"⚠ No config for: {str(question_text)[:60]}... (skipping)")
                continue

            q_type      = config['type']
            options     = config['options']
            rank_labels = config.get('rank_labels', [])
            series      = self.merge_unnamed_columns(q_info)

            is_matrix  = False
            is_bipolar = False

            if q_type == 'bipolar':
                pole_pairs = config.get('pole_pairs', [])
                data       = self.process_bipolar(q_info, options, pole_pairs)
                is_bipolar = True

            elif q_type in ['single', 'single-select']:
                data = self.process_single_select(series, options)

            elif q_type in ['multiple', 'multi-select', 'multiple select']:
                data = self.process_multi_select(series, options)

            elif 'matrix' in q_type:
                data      = self.process_matrix(q_info, options, rank_labels)
                is_matrix = True

            else:
                print(f"⚠ Unknown type '{q_type}' for: {question_text[:60]}... (treating as single)")
                data = self.process_single_select(series, options)

            if not data:
                continue

            if is_matrix:
                response_count = data.get('base', 0)
            elif is_bipolar:
                response_count = data.get('base', 0)
            else:
                response_count = data[0]['base'] if data else 0

            col_index           = q_info['main_col_idx']
            raw_data_col_letter = self.get_excel_column_letter(col_index + 1)

            analysis['questions'].append({
                'question_number':     q_num,
                'question_text':       question_text,
                'question_type':       q_type,
                'is_matrix':           is_matrix,
                'is_bipolar':          is_bipolar,
                'response_count':      response_count,
                'data':                data,
                'raw_data_col_letter': raw_data_col_letter,
                'raw_data_col_index':  col_index
            })

        return analysis

    # ------------------------------------------------------------------
    # CREATE DATABOOK
    # ------------------------------------------------------------------

    def create_databook(self, output_path, survey_name="Survey Databook"):
        analysis = self.analyze_survey()

        wb       = Workbook()
        ws_raw   = wb.active
        ws_raw.title = "Raw Data"
        ws_input  = wb.create_sheet("Input")
        ws_output = wb.create_sheet("Output")
        ws        = ws_output

        header_fill   = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font   = Font(bold=True, color="FFFFFF", size=11)
        question_font = Font(bold=True, size=12, color="1F4E78")
        zero_fill     = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
        border        = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'),  bottom=Side(style='thin')
        )

        ws.merge_cells('A1:D1')
        title_cell       = ws['A1']
        title_cell.value = survey_name
        title_cell.font  = Font(bold=True, size=16, color="1F4E78")
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30

        ws['A3'] = "Total Responses: "
        ws['B3'] = "=COUNTA('Raw Data'!$A$3:$A$8000)"
        ws['A3'].font = Font(bold=True, size=11)
        ws['B3'].font = Font(bold=True, size=11)
        ws['A4'] = f"Total Questions: {len(analysis['questions'])}"
        ws['A4'].font = Font(bold=True, size=11)

        row = 6

        for question in analysis['questions']:
            is_matrix  = question.get('is_matrix', False)
            is_bipolar = question.get('is_bipolar', False)

            merge_to = 'G' if (is_matrix or is_bipolar) else 'D'
            ws.merge_cells(f'A{row}:{merge_to}{row}')
            cell       = ws[f'A{row}']
            cell.value = f"Q{question['question_number']}. {question['question_text']}"
            cell.font  = question_font
            cell.fill  = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            ws.row_dimensions[row].height = 30
            row += 1

            ws[f'A{row}'] = f"Type: {question['question_type'].title()} | Base: "
            ws[f'B{row}'] = "=COUNTA('Raw Data'!$A$3:$A$8000)"
            ws[f'A{row}'].font = Font(italic=True, size=9, color="666666")
            ws[f'B{row}'].font = Font(italic=True, size=9, color="666666")
            row += 1

            # ── BIPOLAR TABLE ──────────────────────────────────────────────
            if is_bipolar:
                data         = question['data']
                bipolar_rows = data['rows']

                first_row  = bipolar_rows[0] if bipolar_rows else {}
                hdr_pole_1 = first_row.get('pole_1', 'Pole 1')
                hdr_pole_2 = first_row.get('pole_2', 'Pole 2')

                headers = ['Response Option', hdr_pole_1, hdr_pole_2, 'N', hdr_pole_1, hdr_pole_2]

                for col_idx, hdr in enumerate(headers, 1):
                    cell       = ws.cell(row, col_idx, hdr)
                    cell.fill  = header_fill
                    cell.font  = header_font
                    cell.alignment = Alignment(horizontal='center')
                    cell.border    = border
                row += 1

                for brow in bipolar_rows:
                    opt         = brow['option']
                    pole_1      = brow['pole_1']
                    pole_2      = brow['pole_2']
                    col_idx_raw = brow['col_idx']

                    cell           = ws.cell(row, 1, opt)
                    cell.border    = border
                    cell.alignment = Alignment(horizontal='left', vertical='center')

                    if col_idx_raw is not None:
                        raw_col_letter = self.get_excel_column_letter(col_idx_raw + 1)

                        p1_formula = (
                            f"=COUNTIFS('Raw Data'!${raw_col_letter}$3"
                            f":${raw_col_letter}$8000,\"{pole_1}\")"
                        )
                        cell           = ws.cell(row, 2, p1_formula)
                        cell.border    = border
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.font      = Font(name="Arial")

                        p2_formula = (
                            f"=COUNTIFS('Raw Data'!${raw_col_letter}$3"
                            f":${raw_col_letter}$8000,\"{pole_2}\")"
                        )
                        cell           = ws.cell(row, 3, p2_formula)
                        cell.border    = border
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.font      = Font(name="Arial")

                        n_formula = (
                            f"=COUNTA('Raw Data'!${raw_col_letter}$3"
                            f":${raw_col_letter}$8000)"
                        )
                        cell           = ws.cell(row, 4, n_formula)
                        cell.border    = border
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.font      = Font(bold=True, name="Arial")
                    else:
                        for c in range(2, 4):
                            cell           = ws.cell(row, c, 0)
                            cell.border    = border
                            cell.fill      = zero_fill
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                            cell.font      = Font(name="Arial")
                        cell           = ws.cell(row, 4, 0)
                        cell.border    = border
                        cell.fill      = zero_fill
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.font      = Font(bold=True, name="Arial")

                    pct1_formula   = f"=IFERROR(B{row}/D{row}*100,0)"
                    cell           = ws.cell(row, 5, pct1_formula)
                    cell.number_format = '0.0"%"'
                    cell.border    = border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.font      = Font(name="Arial")

                    pct2_formula   = f"=IFERROR(C{row}/D{row}*100,0)"
                    cell           = ws.cell(row, 6, pct2_formula)
                    cell.number_format = '0.0"%"'
                    cell.border    = border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.font      = Font(name="Arial")

                    row += 1

                ws.column_dimensions['A'].width = 45
                for col_letter in ['B', 'C', 'D', 'E', 'F']:
                    ws.column_dimensions[col_letter].width = 18

            # ── MATRIX TABLE ───────────────────────────────────────────────
            elif is_matrix:
                data        = question['data']
                rank_labels = data['rank_labels']

                headers        = ['Response Option'] + rank_labels + ['N'] + rank_labels
                header_row_num = row

                for col_idx, header in enumerate(headers, 1):
                    cell       = ws.cell(row, col_idx, header)
                    cell.fill  = header_fill
                    cell.font  = header_font
                    cell.alignment = Alignment(horizontal='center')
                    cell.border    = border
                row += 1

                column_mapping         = data.get('column_mapping', [])
                is_multi_select_matrix = data.get('is_multi_select_matrix', False)

                for item in data['data']:
                    attribute = item['attribute']
                    cell      = ws.cell(row, 1, attribute)
                    cell.border = border

                    for col_idx, rank_label in enumerate(rank_labels, 2):
                        header_col_letter = self.get_excel_column_letter(col_idx)
                        header_cell_ref   = f"${header_col_letter}${header_row_num}"

                        matching_cols = [
                            cm['col_idx'] for cm in column_mapping
                            if cm['attribute'] == attribute and
                            (is_multi_select_matrix and cm['rank_label'] == rank_label
                             or not is_multi_select_matrix)
                        ]

                        formulas = []
                        for raw_col_idx in matching_cols:
                            raw_col_letter = self.get_excel_column_letter(raw_col_idx + 1)
                            formulas.append(
                                f"COUNTIFS('Raw Data'!${raw_col_letter}$3"
                                f":${raw_col_letter}$8000,{header_cell_ref})"
                            )

                        if formulas:
                            cell       = ws.cell(row, col_idx, "=" + "+".join(formulas))
                        else:
                            cell       = ws.cell(row, col_idx, 0)
                            cell.fill  = zero_fill

                        cell.border    = border
                        cell.alignment = Alignment(horizontal='center')

                    n_col_idx     = len(rank_labels) + 2
                    matching_cols = [
                        cm['col_idx'] for cm in column_mapping
                        if cm['attribute'] == attribute
                    ]
                    len_parts = [
                        f"LEN('Raw Data'!${self.get_excel_column_letter(ci + 1)}$3"
                        f":${self.get_excel_column_letter(ci + 1)}$8000)"
                        for ci in matching_cols
                    ]
                    cell       = ws.cell(row, n_col_idx,
                                        f"=SUMPRODUCT(({'+'.join(len_parts)}>0)*1)")
                    cell.border    = border
                    cell.alignment = Alignment(horizontal='center')

                    for rank_idx in range(len(rank_labels)):
                        pct_col_idx    = n_col_idx + 1 + rank_idx
                        rank_count_col = self.get_excel_column_letter(2 + rank_idx)
                        n_col_letter   = self.get_excel_column_letter(n_col_idx)
                        cell           = ws.cell(row, pct_col_idx,
                                                 f"=IFERROR({rank_count_col}{row}/{n_col_letter}{row}*100,0)")
                        cell.number_format = '0.0"%"'
                        cell.border        = border
                        cell.alignment     = Alignment(horizontal='center')

                    row += 1

                ws.column_dimensions['A'].width = 60
                total_cols = len(rank_labels) * 2 + 2
                for col_idx in range(2, total_cols + 1):
                    ws.column_dimensions[self.get_excel_column_letter(col_idx)].width = 12

            # ── SINGLE / MULTIPLE TABLE ────────────────────────────────────
            else:
                for col_idx, header in enumerate(['Response Option', 'N', '%'], 1):
                    cell       = ws.cell(row, col_idx, header)
                    cell.fill  = header_fill
                    cell.font  = header_font
                    cell.alignment = Alignment(horizontal='center')
                    cell.border    = border
                row += 1

                raw_data_col = question.get('raw_data_col_letter', 'A')
                is_multiple  = question['question_type'] in ['multiple', 'multi-select', 'multiple select']

                if is_multiple:
                    q_num        = question['question_number']
                    q_info       = self.question_map[q_num]
                    main_col     = question['raw_data_col_index']
                    unnamed_cols = q_info['unnamed_cols']
                    all_cols     = [main_col] + unnamed_cols
                    all_col_letters = [self.get_excel_column_letter(ci + 1) for ci in all_cols]

                for idx, item in enumerate(question['data']):
                    current_row = row + idx
                    cell        = ws.cell(current_row, 1, item['option'])
                    cell.border = border

                    is_other = item['option'].strip().lower() in (
                        'other (please specify)', 'others (please specify)'
                    )

                    if is_multiple and idx < len(all_col_letters):
                        option_col = all_col_letters[idx]

                        if is_other:
                            next_col_idx    = all_cols[idx] + 1
                            next_col_letter = self.get_excel_column_letter(next_col_idx + 1)
                            count_formula   = (
                                f"=COUNTA('Raw Data'!${next_col_letter}$3"
                                f":${next_col_letter}$8000)"
                            )
                        else:
                            count_formula = (
                                f"=COUNTIFS('Raw Data'!${option_col}$3"
                                f":${option_col}$8000,A{current_row})"
                            )
                    else:
                        if is_other:
                            other_col_idx    = question['raw_data_col_index'] + 1
                            other_col_letter = self.get_excel_column_letter(other_col_idx + 1)
                            count_formula    = (
                                f"=COUNTA('Raw Data'!${other_col_letter}$3"
                                f":${other_col_letter}$8000)"
                            )
                        else:
                            count_formula = (
                                f"=COUNTIFS('Raw Data'!${raw_data_col}$3"
                                f":${raw_data_col}$8000,A{current_row})"
                            )

                    cell        = ws.cell(current_row, 2, count_formula)
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center')

                    n_row_ref   = row + len(question['data'])
                    pct_formula = f"=IFERROR(B{current_row}/B${n_row_ref}*100,0)"
                    cell        = ws.cell(current_row, 3, pct_formula)
                    cell.number_format = '0.0"%"'
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center')

                row += len(question['data'])

                cell        = ws.cell(row, 1, "N")
                cell.border = border
                cell.font   = Font(bold=True)

                first_opt = row - len(question['data'])
                last_opt  = row - 1
                is_single = question['question_type'] in ['single', 'single-select']

                if is_single:
                    n_formula = f"=SUM(B{first_opt}:B{last_opt})"
                    cell      = ws.cell(row, 2, n_formula)
                else:
                    q_num        = question['question_number']
                    q_info       = self.question_map[q_num]
                    main_col     = question['raw_data_col_index']
                    unnamed_cols = q_info['unnamed_cols']
                    all_cols     = [main_col] + unnamed_cols
                    len_parts    = [
                        f"LEN('Raw Data'!{self.get_excel_column_letter(ci + 1)}3"
                        f":{self.get_excel_column_letter(ci + 1)}8000)"
                        for ci in all_cols
                    ]
                    n_formula = (
                        f"=SUMPRODUCT((LEN('Raw Data'!$A$3:$A$8000)>0)"
                        f"*({'+'.join(len_parts)}>0)*1)"
                    )
                    cell = ws.cell(row, 2, n_formula)

                cell.border    = border
                cell.alignment = Alignment(horizontal='center')
                cell.font      = Font(bold=True)

                row += 1

            row += 2

        ws.column_dimensions['A'].width = 70
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15

        # RAW DATA SHEET
        print(f"\n📋 Populating Raw Data sheet...")
        for row_idx in range(len(self.df_full)):
            for col_idx in range(len(self.df_full.columns)):
                val  = self.df_full.iloc[row_idx, col_idx]
                cell = ws_raw.cell(row_idx + 1, col_idx + 1, val)
                if row_idx == 0:
                    cell.font = Font(bold=True)
        print(f"   ✅ {len(self.df_full)} rows × {len(self.df_full.columns)} cols")

        # INPUT SHEET
        print(f"📋 Populating Input sheet...")
        df_opts = pd.read_excel(self.options_file_path)
        for col_idx, col_name in enumerate(df_opts.columns, 1):
            c      = ws_input.cell(1, col_idx, col_name)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        for row_idx, row_val in enumerate(df_opts.itertuples(index=False), 2):
            for col_idx, val in enumerate(row_val, 1):
                ws_input.cell(row_idx, col_idx, val)
        ws_input.column_dimensions['A'].width = 80
        ws_input.column_dimensions['B'].width = 60
        ws_input.column_dimensions['C'].width = 15
        print(f"   ✅ {len(df_opts)} rows")

        wb.save(output_path)
        print(f"\n✅ Databook saved: {output_path}")
        print(f"   {len(analysis['questions'])} questions | {analysis['total_responses']} responses")

        return analysis


# =============================================================================
# CUTS AUTOMATION
# =============================================================================

def load_cuts_from_template(cuts_path):
    wb   = load_workbook(cuts_path, data_only=True)
    ws   = wb.active
    cuts = []

    for row in range(1, 9):
        label       = ws.cell(row=row, column=1).value
        raw_col_val = ws.cell(row=row, column=2).value

        if raw_col_val in (None, ""):
            continue
        try:
            raw_col_index = int(raw_col_val)
        except ValueError:
            continue

        categories = []
        for col in range(4, 4 + 15):
            val = ws.cell(row=row, column=col).value
            if val in (None, ""):
                break
            categories.append(str(val).strip())

        if not categories:
            continue

        cut_index = None
        if isinstance(label, str):
            m = re.search(r"Cut\s*#(\d+)", label)
            if m:
                cut_index = int(m.group(1))

        cuts.append({
            "index":         cut_index,
            "raw_col_index": raw_col_index,
            "categories":    categories,
        })

    if not cuts:
        print("⚠️ No valid cuts found in template")
    else:
        print(f"✅ Loaded {len(cuts)} cuts")
        for c in cuts:
            print(f"   Cut #{c['index']} → col {c['raw_col_index']}, {len(c['categories'])} categories")

    return cuts


def add_demographic_filter_to_formula(formula, cut_raw_col_letter, header_cell_ref):
    if not formula:
        return formula
    f_str = str(formula)
    if f_str.startswith("="):
        f_str_body = f_str[1:]
    else:
        f_str_body = f_str
    if not f_str_body.endswith(")"):
        return formula
    extra = (
        f", 'Raw Data'!${cut_raw_col_letter}$3:${cut_raw_col_letter}$8000,"
        f"{header_cell_ref})"
    )
    return "=" + f_str_body[:-1] + extra


def modify_multiple_select_n_formula(base_formula, cut_raw_col_letter, header_cell_ref):
    if not base_formula:
        return base_formula
    f_str = str(base_formula).strip()
    if not f_str.upper().startswith("=SUMPRODUCT"):
        return base_formula
    if not f_str.endswith("*1)"):
        return base_formula
    formula_body       = f_str[:-3]
    demographic_filter = (
        f"*('Raw Data'!${cut_raw_col_letter}$3:${cut_raw_col_letter}$8000={header_cell_ref})"
    )
    return formula_body + demographic_filter + "*1)"


def find_question_blocks(ws):
    blocks  = []
    max_row = ws.max_row
    r       = 6

    while r <= max_row:
        cell_val = ws.cell(row=r, column=1).value

        if isinstance(cell_val, str) and cell_val.strip().startswith("Q"):
            type_text = str(ws.cell(row=r + 1, column=1).value or "")
            m         = re.search(r"Type:\s*([A-Za-z\s\-\(\)]+)\s*\|", type_text)
            q_type    = m.group(1).strip().lower() if m else ""

            header_row = r + 2
            header_val = ws.cell(row=header_row, column=1).value

            is_matrix        = "matrix" in q_type
            is_bipolar       = "bipolar" in q_type
            num_rating_cols  = 0

            if header_val == "Response Option":
                if is_matrix:
                    col_idx = 2
                    while col_idx <= 200:
                        if ws.cell(row=header_row, column=col_idx).value == "N":
                            break
                        num_rating_cols += 1
                        col_idx         += 1
                elif is_bipolar:
                    num_rating_cols = 2
            else:
                r += 1
                continue

            cur           = header_row + 1
            n_row         = None
            last_data_row = None

            while cur <= max_row:
                a_val = ws.cell(row=cur, column=1).value

                if a_val == "N" and not is_matrix and not is_bipolar:
                    n_row = cur
                    break

                if a_val is None and ws.cell(row=cur, column=2).value is None:
                    if (is_matrix or is_bipolar) and last_data_row:
                        n_row = last_data_row
                    break

                if isinstance(a_val, str) and re.match(r'^Q\d+', a_val.strip()):
                    if (is_matrix or is_bipolar) and last_data_row:
                        n_row = last_data_row
                    break

                if a_val and isinstance(a_val, str):
                    last_data_row = cur

                cur += 1

            if not n_row:
                r += 1
                continue

            first_option_row = header_row + 1
            last_option_row  = n_row if (is_matrix or is_bipolar) else n_row - 1

            blocks.append({
                "header_row":       header_row,
                "first_option_row": first_option_row,
                "last_option_row":  last_option_row,
                "n_row":            n_row,
                "q_type":           q_type,
                "is_matrix":        is_matrix,
                "is_bipolar":       is_bipolar,
                "num_rating_cols":  num_rating_cols,
            })
            r = n_row + 3
        else:
            r += 1

    matrix_count  = sum(1 for b in blocks if b.get('is_matrix'))
    bipolar_count = sum(1 for b in blocks if b.get('is_bipolar'))
    print(
        f"✅ Found {len(blocks)} blocks "
        f"({matrix_count} matrix, {bipolar_count} bipolar, "
        f"{len(blocks) - matrix_count - bipolar_count} single/multiple)"
    )
    return blocks


def apply_cuts_to_databook(databook_path, cuts_template_path):
    cuts = load_cuts_from_template(cuts_template_path)
    if not cuts:
        return

    wb = load_workbook(databook_path)
    if "Output" not in wb.sheetnames:
        raise ValueError("Sheet 'Output' not found.")
    ws = wb["Output"]

    max_question_width = 3
    r = 6
    while r <= ws.max_row:
        cell_val = ws.cell(row=r, column=1).value
        if isinstance(cell_val, str) and cell_val.strip().startswith("Q"):
            type_text  = str(ws.cell(row=r + 1, column=1).value or "")
            is_matrix  = "matrix"  in type_text.lower()
            is_bipolar = "bipolar" in type_text.lower()

            if is_matrix or is_bipolar:
                header_row = r + 2
                col_idx    = 1
                while col_idx <= 200:
                    if ws.cell(row=header_row, column=col_idx).value is None:
                        break
                    col_idx += 1
                max_question_width = max(max_question_width, col_idx - 1)
        r += 1

    print(f"📊 Widest base question: {max_question_width} columns")

    cuts_start_col    = max_question_width + 7
    current_start_col = cuts_start_col
    blocks            = find_question_blocks(ws)
    cut_blocks        = []

    grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    for cut_idx, cut in enumerate(cuts, start=1):
        num_cat   = len(cut["categories"])
        start_col = current_start_col

        max_cut_width = 2 * num_cat

        for block in blocks:
            if block.get("is_matrix"):
                nr            = block.get("num_rating_cols", 0)
                max_cut_width = max(max_cut_width, nr + 1 + nr)
            elif block.get("is_bipolar"):
                max_cut_width = max(max_cut_width, 5)

        end_col = start_col + max_cut_width - 1

        cut_blocks.append({
            "raw_col_index":  cut["raw_col_index"],
            "raw_col_letter": get_column_letter(cut["raw_col_index"]),
            "categories":     cut["categories"],
            "num_cat":        num_cat,
            "start_col":      start_col,
            "end_col":        end_col,
            "max_width":      max_cut_width,
        })

        current_start_col = end_col + 7

    for cb in cut_blocks:
        start_col  = cb["start_col"]
        grey_col_1 = start_col - 4
        grey_col_2 = start_col - 3

        for row_idx in range(1, ws.max_row + 1):
            ws.cell(row=row_idx, column=grey_col_1).fill = grey_fill
            ws.cell(row=row_idx, column=grey_col_2).fill = grey_fill

        dropdown_cell       = ws.cell(row=1, column=start_col)
        all_options         = ["<>"] + cb["categories"]
        dropdown_cell.value = "<>"
        dv = DataValidation(type="list", formula1=f'"{",".join(all_options)}"', allow_blank=False)
        dv.add(dropdown_cell)
        ws.add_data_validation(dv)
        dropdown_cell.font      = Font(bold=True, size=11)
        dropdown_cell.fill      = PatternFill(start_color="90D9D6", end_color="90D9D6", fill_type="solid")
        dropdown_cell.alignment = Alignment(horizontal='center', vertical='center')

    for q_idx, block in enumerate(blocks, start=1):
        q_type     = block["q_type"]
        is_matrix  = block.get("is_matrix", False)
        is_bipolar = block.get("is_bipolar", False)
        header_row = block["header_row"]
        first_opt  = block["first_option_row"]
        last_opt   = block["last_option_row"]
        n_row      = block["n_row"]

        if is_bipolar:
            print(f"\n➡️ Bipolar cuts for block #{q_idx} (rows {header_row}-{last_opt})")

            for cb in cut_blocks:
                start_col          = cb["start_col"]
                cut_raw_col_letter = cb["raw_col_letter"]

                dropdown_col_letter = get_column_letter(start_col)
                dropdown_ref        = f"${dropdown_col_letter}$1"

                for offset in range(5):
                    src_cell  = ws.cell(row=header_row, column=2 + offset)
                    dest_cell = ws.cell(row=header_row, column=start_col + offset)
                    dest_cell.value = src_cell.value
                    if src_cell.has_style:
                        dest_cell._style = copy(src_cell._style)

                for data_row in range(first_opt, last_opt + 1):
                    base_p1 = ws.cell(row=data_row, column=2).value
                    base_p2 = ws.cell(row=data_row, column=3).value
                    base_n  = ws.cell(row=data_row, column=4).value

                    mod_p1 = (
                        add_demographic_filter_to_formula(base_p1, cut_raw_col_letter, dropdown_ref)
                        if isinstance(base_p1, str) and base_p1.startswith("=") else base_p1
                    )
                    mod_p2 = (
                        add_demographic_filter_to_formula(base_p2, cut_raw_col_letter, dropdown_ref)
                        if isinstance(base_p2, str) and base_p2.startswith("=") else base_p2
                    )

                    if isinstance(base_n, str) and base_n.startswith("=COUNTA"):
                        m = re.search(
                            r"COUNTA\('Raw Data'!\$([A-Z]+)\$3:\$([A-Z]+)\$8000\)", base_n
                        )
                        if m:
                            raw_col = m.group(1)
                            mod_n   = (
                                f"=COUNTIFS('Raw Data'!${raw_col}$3:${raw_col}$8000,\"<>\","
                                f"'Raw Data'!${cut_raw_col_letter}$3:${cut_raw_col_letter}$8000,"
                                f"{dropdown_ref})"
                            )
                        else:
                            mod_n = base_n
                    else:
                        mod_n = base_n

                    border_thin = Border(
                        left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'),  bottom=Side(style='thin')
                    )
                    center = Alignment(horizontal='center', vertical='center')

                    c = ws.cell(row=data_row, column=start_col, value=mod_p1)
                    c.border = border_thin; c.alignment = center; c.font = Font(name="Arial")

                    c = ws.cell(row=data_row, column=start_col + 1, value=mod_p2)
                    c.border = border_thin; c.alignment = center; c.font = Font(name="Arial")

                    c = ws.cell(row=data_row, column=start_col + 2, value=mod_n)
                    c.border = border_thin; c.alignment = center; c.font = Font(bold=True, name="Arial")

                    p1_col = get_column_letter(start_col)
                    p2_col = get_column_letter(start_col + 1)
                    n_col  = get_column_letter(start_col + 2)

                    pct1 = ws.cell(
                        row=data_row, column=start_col + 3,
                        value=f"=IFERROR({p1_col}{data_row}/{n_col}{data_row}*100,0)"
                    )
                    pct1.number_format = '0.0"%"'; pct1.border = border_thin
                    pct1.alignment = center; pct1.font = Font(name="Arial")

                    pct2 = ws.cell(
                        row=data_row, column=start_col + 4,
                        value=f"=IFERROR({p2_col}{data_row}/{n_col}{data_row}*100,0)"
                    )
                    pct2.number_format = '0.0"%"'; pct2.border = border_thin
                    pct2.alignment = center; pct2.font = Font(name="Arial")

            print(f"   ✓ Bipolar cuts applied")

        elif is_matrix:
            num_rating_cols = block.get("num_rating_cols", 0)
            print(f"\n➡️ Matrix cuts for block #{q_idx}")

            for cb in cut_blocks:
                start_col          = cb["start_col"]
                cut_raw_col_letter = cb["raw_col_letter"]
                total_matrix_cols  = num_rating_cols + 1 + num_rating_cols

                for row_offset in range(last_opt - header_row + 1):
                    src_row = header_row + row_offset
                    for col_offset in range(total_matrix_cols):
                        src_col   = 2 + col_offset
                        dest_col  = start_col + col_offset
                        src_cell  = ws.cell(row=src_row, column=src_col)
                        dest_cell = ws.cell(row=src_row, column=dest_col)
                        dest_cell.value = src_cell.value
                        if src_cell.has_style:
                            dest_cell._style = copy(src_cell._style)
                        if src_cell.number_format:
                            dest_cell.number_format = src_cell.number_format

                dropdown_col_letter = get_column_letter(start_col)
                dropdown_ref        = f"${dropdown_col_letter}$1"

                for rating_offset in range(num_rating_cols):
                    rating_col = start_col + rating_offset
                    for data_row in range(first_opt, last_opt + 1):
                        cell    = ws.cell(row=data_row, column=rating_col)
                        formula = cell.value
                        if isinstance(formula, str) and formula.startswith("=") and formula.endswith(")"):
                            cut_cond   = (
                                f",'Raw Data'!${cut_raw_col_letter}$3"
                                f":${cut_raw_col_letter}$8000,{dropdown_ref})"
                            )
                            cell.value = formula[:-1] + cut_cond

                base_n_col_pos = 1 + num_rating_cols + 1
                n_col_idx      = start_col + num_rating_cols

                for data_row in range(first_opt, last_opt + 1):
                    base_formula = str(ws.cell(row=data_row, column=base_n_col_pos).value or "")
                    if 'SUMPRODUCT' in base_formula.upper() and base_formula.endswith("*1)"):
                        formula_body       = base_formula[:-3]
                        demographic_filter = (
                            f"*('Raw Data'!${cut_raw_col_letter}$3"
                            f":${cut_raw_col_letter}$8000={dropdown_ref})"
                        )
                        ws.cell(row=data_row, column=n_col_idx).value = (
                            formula_body + demographic_filter + "*1)"
                        )

        elif "single" in q_type or "multiple" in q_type or "multi-select" in q_type:
            print(f"\n➡️ Single/Multiple cuts for block #{q_idx}")

            for cb in cut_blocks:
                start_col          = cb["start_col"]
                num_cat            = cb["num_cat"]
                categories         = cb["categories"]
                cut_raw_col_letter = cb["raw_col_letter"]

                for i, cat in enumerate(categories):
                    n_hdr   = ws.cell(row=header_row, column=start_col + i)
                    pct_hdr = ws.cell(row=header_row, column=start_col + num_cat + i)
                    n_hdr.value   = cat
                    pct_hdr.value = cat
                    base_hdr = ws.cell(row=header_row, column=2)
                    if base_hdr.has_style:
                        n_hdr._style   = copy(base_hdr._style)
                        pct_hdr._style = copy(base_hdr._style)

                for data_row in range(first_opt, last_opt + 1):
                    base_formula = ws.cell(row=data_row, column=2).value

                    if not (isinstance(base_formula, str) and base_formula.startswith("=")):
                        continue

                    is_counta = base_formula.upper().startswith("=COUNTA")

                    for i, cat in enumerate(categories):
                        n_col           = start_col + i
                        hdr_col_letter  = get_column_letter(n_col)
                        header_cell_ref = f"${hdr_col_letter}${header_row}"

                        if is_counta:
                            m = re.search(
                                r"COUNTA\('Raw Data'!\$([A-Z]+)\$3:\$([A-Z]+)\$8000\)",
                                base_formula
                            )
                            if m:
                                other_col        = m.group(1)
                                modified_formula = (
                                    f"=COUNTIFS('Raw Data'!${other_col}$3:${other_col}$8000,\"<>\")"
                                    f",'Raw Data'!${cut_raw_col_letter}$3:${cut_raw_col_letter}$8000,"
                                    f"{header_cell_ref})"
                                )
                            else:
                                modified_formula = base_formula
                        else:
                            modified_formula = add_demographic_filter_to_formula(
                                base_formula, cut_raw_col_letter, header_cell_ref
                            )

                        n_cell       = ws.cell(row=data_row, column=n_col)
                        n_cell.value = modified_formula
                        base_n_cell  = ws.cell(row=data_row, column=2)
                        if base_n_cell.has_style:
                            n_cell._style = copy(base_n_cell._style)

                        pct_col      = start_col + num_cat + i
                        pct_cell     = ws.cell(row=data_row, column=pct_col)
                        pct_cell.value = (
                            f"=IFERROR({hdr_col_letter}{data_row}"
                            f"/{hdr_col_letter}{n_row}*100,0)"
                        )
                        pct_cell.number_format = '0.0"%"'
                        base_pct = ws.cell(row=data_row, column=3)
                        if base_pct.has_style:
                            pct_cell._style = copy(base_pct._style)

                for i, cat in enumerate(categories):
                    n_col          = start_col + i
                    n_col_letter   = get_column_letter(n_col)
                    base_total     = ws.cell(row=n_row, column=2)
                    base_n_formula = base_total.value
                    total_cell     = ws.cell(row=n_row, column=n_col)

                    if isinstance(base_n_formula, str) and 'SUMPRODUCT' in base_n_formula.upper():
                        hdr_col_letter  = get_column_letter(n_col)
                        header_cell_ref = f"${hdr_col_letter}${header_row}"
                        total_cell.value = modify_multiple_select_n_formula(
                            base_n_formula, cut_raw_col_letter, header_cell_ref
                        )
                    else:
                        total_cell.value = f"=SUM({n_col_letter}{first_opt}:{n_col_letter}{last_opt})"

                    if base_total.has_style:
                        total_cell._style = copy(base_total._style)

    wb.save(databook_path)
    print(f"\n✅ Cuts applied → '{databook_path}'")


# =============================================================================
# ███╗   ███╗ █████╗ ██╗███╗   ██╗
# ████╗ ████║██╔══██╗██║████╗  ██║
# ██╔████╔██║███████║██║██╔██╗ ██║
# ██║╚██╔╝██║██╔══██║██║██║╚██╗██║
# ██║ ╚═╝ ██║██║  ██║██║██║ ╚████║
# ╚═╝     ╚═╝╚═╝  ╚═╝╚═╝╚═╝  ╚═══╝
#
# Just update the 3 paths below, then run this file.
# Cuts are configured interactively in the terminal — no template file needed.
# =============================================================================


def create_cuts_template_from_config(cuts_config):
    """
    Converts the terminal-collected cuts_config list into a temporary
    Template_Cuts.xlsx workbook in memory — exactly the same structure
    that load_cuts_from_template() expects.

    cuts_config format:
        [
            {'index': 1, 'raw_col_index': 5, 'categories': ['Male', 'Female']},
            {'index': 2, 'raw_col_index': 8, 'categories': ['18-24', '25-34']},
        ]
    """
    wb = Workbook()
    ws = wb.active

    for cut in cuts_config:
        row = cut['index']
        ws.cell(row=row, column=1, value=f"Cut #{cut['index']} column =")
        ws.cell(row=row, column=2, value=cut['raw_col_index'])
        ws.cell(row=row, column=3, value=f"Cut #{cut['index']} parameter =")
        for cat_idx, category in enumerate(cut['categories'], start=4):
            ws.cell(row=row, column=cat_idx, value=category)

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(tmp.name)
    tmp.close()
    return tmp.name


def collect_cuts_from_terminal(data_file):
    """
    Interactively prompts the user in the terminal to configure cuts.
    Shows available column names from the raw data file to help the user.
    Returns a cuts_config list ready for create_cuts_template_from_config().
    """
    print("\n" + "─" * 60)
    print("✂️  CUTS CONFIGURATION")
    print("─" * 60)

    # Show available columns
    if data_file.endswith('.csv'):
        df_cols = pd.read_csv(data_file, nrows=0)
    else:
        df_cols = pd.read_excel(data_file, nrows=0)

    print("\n📋 Available columns in your raw data:")
    for i, col in enumerate(df_cols.columns, 1):
        print(f"   Col {i:>3} : {col}")

    print()

    # Ask how many cuts
    while True:
        try:
            num_cuts = int(input("How many demographic cuts do you want? (0-8): ").strip())
            if 0 <= num_cuts <= 8:
                break
            else:
                print("❌ Please enter a number between 0 and 8.")
        except ValueError:
            print("❌ Invalid input. Please enter a number.")

    if num_cuts == 0:
        print("✅ No cuts selected — databook will be generated without demographic filters.")
        return []

    cuts_config = []

    for i in range(1, num_cuts + 1):
        print(f"\n--- Cut #{i} ---")

        # Column number
        while True:
            try:
                col_num = int(input(f"  Enter column number for Cut #{i}: ").strip())
                if 1 <= col_num <= len(df_cols.columns):
                    col_name = df_cols.columns[col_num - 1]
                    print(f"  📌 Selected: '{col_name}'")
                    break
                else:
                    print(f"❌ Please enter a number between 1 and {len(df_cols.columns)}.")
            except ValueError:
                print("❌ Invalid input. Please enter a number.")

        # Categories
        while True:
            cats_input = input(f"  Enter categories for Cut #{i} (comma-separated): ").strip()
            categories = [c.strip() for c in cats_input.split(',') if c.strip()]
            if categories:
                print(f"  ✅ {len(categories)} categories: {categories}")
                break
            else:
                print("❌ Please enter at least one category.")

        cuts_config.append({
            'index':         i,
            'raw_col_index': col_num,
            'categories':    categories,
        })

    print(f"\n✅ {len(cuts_config)} cut(s) configured successfully.")
    return cuts_config

def main():
    # ──────────────────────────────────────────────────────────────────────────
    # ✏️  UPDATE THESE 3 PATHS ONLY — cuts are configured interactively below
    # ──────────────────────────────────────────────────────────────────────────

    # 1. Raw survey output Excel (input for Step 0 — clean & convert)
    raw_survey_file  = r"C:\Users\1Lattice User\Downloads\Copy of Jewellery Consumer study.xlsx"

    # 2. Intermediate options file (auto-created by Step 0, auto-used by Step 1)
    options_file     = r"C:\Users\1Lattice User\Downloads\Clean Jewellery Consumer study.xlsx"

    # 3. Raw respondent-level data Excel (input for Step 1 databook)
    data_file        = r"C:\Users\1Lattice User\OneDrive - Lattice Technologies Private Limited\Desktop\Raw data-Jewellery.xlsx"

    # 4. Final databook output path
    output_file      = r"C:\Users\1Lattice User\Downloads\Databook Jewellery Consumer study.xlsx"

    # Survey name shown in the databook title
    survey_name      = "Survey Analysis"

    # ──────────────────────────────────────────────────────────────────────────

    print("=" * 80)
    print("SURVEY DATABOOK GENERATOR — All-in-One (Step 0 + Step 1 + Step 2)")
    print("=" * 80)

    try:
        # ── STEP 0 (auto): Clean & convert raw survey → options file ──────────
        print("\n── STEP 0: Clean & convert raw survey output ──")
        clean_and_convert(raw_survey_file, options_file)

        # ── STEP 1: Build base databook ───────────────────────────────────────
        print("\n── STEP 1: Build base databook ──")
        processor = SurveyDatabookV2(data_file, options_file)
        processor.load_options_file()
        processor.load_data()
        processor.identify_structure()
        processor.create_databook(output_file, survey_name)

        # ── STEP 2: Collect cuts interactively, then apply ────────────────────
        print("\n── STEP 2: Configure & apply cuts ──")
        cuts_config = collect_cuts_from_terminal(data_file)

        if cuts_config:
            cuts_template_path = create_cuts_template_from_config(cuts_config)
            apply_cuts_to_databook(output_file, cuts_template_path)

            # Clean up the temp template file
            os.unlink(cuts_template_path)
        else:
            print("⏭️  Skipping cuts — none configured.")

        print("\n" + "=" * 80)
        print("✅ ALL DONE — Databook complete!")
        print(f"   📁 Saved at: {output_file}")
        print("=" * 80)

    except Exception as e:
        print(f"\n❌ Error: {e}")
        import traceback
        traceback.print_exc()




# =============================================================================
# =============================================================================
#
#  ███████╗ █████╗ ███████╗████████╗ █████╗ ██████╗ ██╗
#  ██╔════╝██╔══██╗██╔════╝╚══██╔══╝██╔══██╗██╔══██╗██║
#  █████╗  ███████║███████╗   ██║   ███████║██████╔╝██║
#  ██╔══╝  ██╔══██║╚════██║   ██║   ██╔══██║██╔═══╝ ██║
#  ██║     ██║  ██║███████║   ██║   ██║  ██║██║     ██║
#  ╚═╝     ╚═╝  ╚═╝╚══════╝   ╚═╝   ╚═╝  ╚═╝╚═╝     ╚═╝
#
#  API LAYER — wraps all the logic above into 3 HTTP endpoints.
#  Nothing above this line was changed.
#
#  Endpoints:
#    POST /upload          → upload files, get session_id + questions
#    POST /confirm-types   → submit confirmed question types
#    POST /generate        → submit cuts, get back .xlsx databook
#    DELETE /session/{id}  → clean up temp files
#
#  Run:
#    pip install fastapi uvicorn python-multipart
#    uvicorn survey_databook_api_v2:app --host 0.0.0.0 --port 8000 --reload
#
# =============================================================================
# =============================================================================

import uuid
import shutil
import httpx

from fastapi import FastAPI, UploadFile, File, HTTPException, Header, Depends
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from pydantic import BaseModel
from typing import List, Optional
from datetime import datetime


# ---------------------------------------------------------------------------
# APP SETUP
# ---------------------------------------------------------------------------

app = FastAPI(
    title="Survey Databook Generator API",
    description=(
        "Converts raw survey files into formatted Excel databooks "
        "with demographic cuts. All processing logic is identical to "
        "the standalone merged script."
    ),
    version="2.0.0",
)

# CORS — allow Lovable (or any frontend) to call this API.
# In production replace allow_origins=["*"] with your Lovable app URL.
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Supabase configuration
SUPABASE_URL = "https://xtxhneblpxnjsuudefoo.supabase.co"
SUPABASE_ANON_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6Inh0eGhuZWJscHhuanN1dWRlZm9vIiwicm9sZSI6ImFub24iLCJpYXQiOjE3NTk3MjQ3MjcsImV4cCI6MjA3NTMwMDcyN30.pkEIYGkxpbGMhmrucVRozo6HvGmF2MjGF6OtaC1ZgRc"

# Shared temp directory — one sub-folder per session, cleaned up on /session DELETE
_TEMP_ROOT = tempfile.mkdtemp(prefix="survey_databook_api_")

# In-memory session store
# session_id → {
#     "user_id":         str,   user ID from Supabase JWT
#     "questions_path":  str,   path to saved dirty-questions file
#     "raw_data_path":   str,   path to saved raw-data file
#     "options_path":    str,   path where options file will be written
#     "questions":       list,  parsed question dicts (from parse_questions)
#     "confirmed_types": dict,  q_text → confirmed type string
#     "columns":         list,  [{index, name, sample_values}] for cuts UI
# }
_SESSIONS: dict = {}


# ---------------------------------------------------------------------------
# AUTHENTICATION
# ---------------------------------------------------------------------------

async def get_current_user(authorization: Optional[str] = Header(None)) -> str:
    """
    Validates the Supabase JWT by calling Supabase's auth API.
    Returns the user_id if valid, raises 401 if invalid.
    This is used as a dependency on all protected endpoints.
    """
    if not authorization:
        raise HTTPException(
            status_code=401,
            detail="Missing Authorization header. Please provide a valid Supabase JWT."
        )
    
    # Extract token from "Bearer <token>" format
    if not authorization.startswith("Bearer "):
        raise HTTPException(
            status_code=401,
            detail="Invalid Authorization header format. Expected 'Bearer <token>'"
        )
    
    token = authorization.replace("Bearer ", "").strip()
    
    # Validate JWT by calling Supabase
    async with httpx.AsyncClient() as client:
        try:
            response = await client.get(
                f"{SUPABASE_URL}/auth/v1/user",
                headers={
                    "Authorization": f"Bearer {token}",
                    "apikey": SUPABASE_ANON_KEY
                },
                timeout=10.0
            )
            
            if response.status_code != 200:
                raise HTTPException(
                    status_code=401,
                    detail="Invalid or expired token. Please log in again."
                )
            
            user_data = response.json()
            user_id = user_data.get("id")
            
            if not user_id:
                raise HTTPException(
                    status_code=401,
                    detail="Token validation failed: no user ID found."
                )
            
            return user_id
            
        except httpx.RequestError as e:
            raise HTTPException(
                status_code=503,
                detail=f"Unable to verify token with Supabase: {str(e)}"
            )
        except Exception as e:
            raise HTTPException(
                status_code=401,
                detail=f"Token validation failed: {str(e)}"
            )


# ---------------------------------------------------------------------------
# PYDANTIC MODELS
# ---------------------------------------------------------------------------

class QuestionTypeItem(BaseModel):
    q_text: str
    q_type: str   # "Single" | "Multiple" | "Matrix" | "Bipolar"

class ConfirmTypesRequest(BaseModel):
    session_id:     str
    question_types: List[QuestionTypeItem]

class CutConfigItem(BaseModel):
    index:         int
    raw_col_index: int
    categories:    List[str]

class GenerateRequest(BaseModel):
    session_id:  str
    cuts:        List[CutConfigItem] = []
    survey_name: str = "Survey Databook"


# ---------------------------------------------------------------------------
# INTERNAL HELPERS — used only by the API layer
# ---------------------------------------------------------------------------

def _parse_questions_from_file(file_path: str) -> list:
    """
    Parses the dirty questions Excel file into a list of question dicts.
    Uses all the same helper functions as clean_and_convert() —
    is_question_row, detect_bipolar_question, detect_horizontal_scale, etc.
    — but returns the parsed list instead of writing a file.
    Each dict: { q_text, options, rank_labels, is_bipolar, auto_type }
    """
    df = pd.read_excel(file_path, header=None)

    questions          = []
    current_q_text     = None
    current_options    = []
    current_rank_labels = []
    current_is_bipolar = False
    in_question        = False

    def _flush():
        if current_q_text and (current_options or current_rank_labels):
            opts = current_options.copy()
            if not current_rank_labels:
                opts = expand_nps_if_needed(current_q_text, opts)
            questions.append({
                "q_text":      current_q_text,
                "options":     opts,
                "rank_labels": current_rank_labels.copy(),
                "is_bipolar":  current_is_bipolar,
            })

    idx = 0
    while idx < len(df):
        first_cell = str(df.iloc[idx, 0]).strip() if pd.notna(df.iloc[idx, 0]) else ""

        # Empty row → flush current question
        if first_cell == "":
            _flush()
            current_q_text = None; current_options = []
            current_rank_labels = []; current_is_bipolar = False; in_question = False
            idx += 1
            continue

        # Question header
        if is_question_row(first_cell):
            _flush()
            _, q_text          = parse_question_row(first_cell)
            current_q_text     = q_text
            current_options    = []
            current_rank_labels = []
            current_is_bipolar = False
            in_question        = True

            # Bipolar check
            bipolar_labels, pole_1, pole_2 = detect_bipolar_question(df, idx)
            if bipolar_labels:
                current_options     = bipolar_labels
                current_rank_labels = [pole_1 or "Pole_1", pole_2 or "Pole_2"]
                current_is_bipolar  = True
                skip_to = idx + 1
                while skip_to < len(df):
                    ahead = str(df.iloc[skip_to, 0]).strip() if pd.notna(df.iloc[skip_to, 0]) else ""
                    if is_question_row(ahead):
                        break
                    skip_to += 1
                idx = skip_to
                continue

            # Matrix check
            if idx + 1 < len(df):
                next_row = df.iloc[idx + 1]
                ranks = []
                for col in range(1, min(30, len(next_row))):
                    val = next_row.iloc[col]
                    if pd.notna(val):
                        val_clean = str(val).strip()
                        if val_clean.lower() == "total":
                            break
                        if not is_metadata_text(val_clean):
                            ranks.append(val_clean)
                if len(ranks) >= 2:
                    attributes = []
                    for r in range(idx + 2, len(df)):
                        v = df.iloc[r, 0]
                        if pd.isna(v):
                            break
                        attr = str(v).strip()
                        if is_question_row(attr):
                            break
                        if not is_metadata_text(attr):
                            attributes.append(attr)
                    if len(attributes) >= 2:
                        current_options     = attributes
                        current_rank_labels = ranks

            # Horizontal scale check
            if not current_rank_labels:
                scale = detect_horizontal_scale(df, idx)
                if scale:
                    current_options = scale

            idx += 1
            continue

        # Inside a question — collect options
        elif in_question and not current_rank_labels:
            if not is_metadata_text(first_cell) and first_cell.lower() != "answer choices":
                current_options.append(first_cell)

        idx += 1

    _flush()  # save last question

    # Attach auto_type
    for q in questions:
        q["auto_type"] = get_auto_type(
            q["q_text"], q["rank_labels"], q["options"], q["is_bipolar"]
        )

    return questions


def _write_options_file(questions: list, confirmed_types: dict, output_path: str):
    """
    Writes the clean Question_Options Excel file.
    Uses confirmed_types dict { q_text → type } to set the Type column.
    Uses the same openpyxl styling as the original clean_and_convert().
    """
    max_ranks = max((len(q["rank_labels"]) for q in questions), default=0)
    headers   = ["Question Text", "Option", "Type"] + [f"Rank_{i}" for i in range(1, max_ranks + 1)]

    wb = Workbook()
    ws = wb.active
    ws.title = "Question_Options"
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF", name="Arial")
        cell.fill = PatternFill(start_color="366092", fill_type="solid")

    row = 2
    for q in questions:
        # Use confirmed type if provided, else auto_type, else default to Single
        q_type = confirmed_types.get(q["q_text"], q.get("auto_type") or "Single")

        for opt in q["options"]:
            ws.cell(row, 1, q["q_text"]).font = Font(name="Arial")
            ws.cell(row, 2, clean_value(opt)).font = Font(name="Arial")
            ws.cell(row, 3, q_type).font = Font(name="Arial")
            for r_idx, rank in enumerate(q["rank_labels"]):
                ws.cell(row, 4 + r_idx, clean_value(rank)).font = Font(name="Arial")
            row += 1

    for col in ws.columns:
        max_len    = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

    wb.save(output_path)


def _build_cuts_template(cuts: List[CutConfigItem]) -> str:
    """
    Converts the API cuts list into a temp Template_Cuts.xlsx.
    Identical logic to create_cuts_template_from_config() in the merged script.
    Returns the temp file path.
    """
    wb = Workbook()
    ws = wb.active
    for cut in cuts:
        row = cut.index
        ws.cell(row=row, column=1, value=f"Cut #{cut.index} column =")
        ws.cell(row=row, column=2, value=cut.raw_col_index)
        ws.cell(row=row, column=3, value=f"Cut #{cut.index} parameter =")
        for cat_idx, category in enumerate(cut.categories, start=4):
            ws.cell(row=row, column=cat_idx, value=category)
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx", dir=_TEMP_ROOT)
    wb.save(tmp.name)
    tmp.close()
    return tmp.name


# ---------------------------------------------------------------------------
# ROOT
# ---------------------------------------------------------------------------

@app.get("/")
def root():
    return {
        "service": "Survey Databook Generator API",
        "version": "2.0.0",
        "endpoints": {
            "POST /upload":         "Upload dirty questions + raw data → session_id + questions list",
            "POST /confirm-types":  "Confirm question types for the session",
            "POST /generate":       "Configure cuts → generate and download databook Excel",
            "DELETE /session/{id}": "Clean up temp files after download",
        },
        "docs": "/docs",
    }


# ---------------------------------------------------------------------------
# ENDPOINT 1 — UPLOAD
# ---------------------------------------------------------------------------

@app.post("/upload")
async def upload_files(
    questions_file: UploadFile = File(..., description="Dirty survey questions Excel file"),
    raw_data_file:  UploadFile = File(..., description="Raw respondent data Excel or CSV"),
    user_id: str = Depends(get_current_user),  # ← Authentication required
):
    """
    Step 1.
    Accepts two file uploads, parses the questions file using the same
    logic as clean_and_convert(), and returns:
      - session_id  (pass this to all subsequent calls)
      - questions   (list with auto-detected types — frontend shows these for review)
      - columns     (list of raw-data columns — frontend shows these for cuts config)
    
    **Requires authentication**: Authorization header with Supabase JWT.
    """
    session_id  = str(uuid.uuid4())
    session_dir = os.path.join(_TEMP_ROOT, session_id)
    os.makedirs(session_dir, exist_ok=True)

    # Determine file extensions
    q_ext  = ".xlsx" if questions_file.filename.lower().endswith((".xlsx", ".xls")) else ".xlsx"
    rd_ext = ".csv"  if raw_data_file.filename.lower().endswith(".csv") else ".xlsx"

    questions_path = os.path.join(session_dir, f"questions{q_ext}")
    raw_data_path  = os.path.join(session_dir, f"raw_data{rd_ext}")
    options_path   = os.path.join(session_dir, "options.xlsx")
    output_path    = os.path.join(session_dir, "databook.xlsx")

    # Save uploaded files to disk
    with open(questions_path, "wb") as f:
        f.write(await questions_file.read())
    with open(raw_data_path, "wb") as f:
        f.write(await raw_data_file.read())

    # Parse questions using the same logic as the merged script
    try:
        questions = _parse_questions_from_file(questions_path)
    except Exception as e:
        shutil.rmtree(session_dir, ignore_errors=True)
        raise HTTPException(status_code=400, detail=f"Failed to parse questions file: {str(e)}")

    # Read column list from raw data (used by cuts configuration UI)
    try:
        if rd_ext == ".csv":
            df_preview = pd.read_csv(raw_data_path, nrows=3)
        else:
            df_preview = pd.read_excel(raw_data_path, nrows=3)

        columns = [
            {
                "index":         i + 1,
                "name":          str(col),
                "sample_values": [str(v) for v in df_preview[col].dropna().tolist()[:3]],
            }
            for i, col in enumerate(df_preview.columns)
        ]
    except Exception as e:
        shutil.rmtree(session_dir, ignore_errors=True)
        raise HTTPException(status_code=400, detail=f"Failed to read raw data file: {str(e)}")

    # Store session with user_id for ownership validation
    _SESSIONS[session_id] = {
        "user_id":         user_id,  # ← Store authenticated user ID
        "questions_path":  questions_path,
        "raw_data_path":   raw_data_path,
        "options_path":    options_path,
        "output_path":     output_path,
        "questions":       questions,
        "confirmed_types": {},
        "columns":         columns,
    }

    # Build response
    auto_detected = sum(1 for q in questions if q["auto_type"])
    needs_review  = len(questions) - auto_detected

    questions_response = [
        {
            "q_text":         q["q_text"],
            "options_preview": q["options"][:5],        # first 5 for display
            "total_options":  len(q["options"]),
            "rank_labels":    q["rank_labels"],
            "is_bipolar":     q["is_bipolar"],
            "auto_type":      q["auto_type"],            # "" = needs manual selection
            "needs_review":   q["auto_type"] == "",
        }
        for q in questions
    ]

    return {
        "session_id":       session_id,
        "total_questions":  len(questions),
        "auto_detected":    auto_detected,
        "needs_review":     needs_review,
        "questions":        questions_response,
        "columns":          columns,
    }


# ---------------------------------------------------------------------------
# ENDPOINT 2 — CONFIRM TYPES
# ---------------------------------------------------------------------------

@app.post("/confirm-types")
async def confirm_types(
    request: ConfirmTypesRequest,
    user_id: str = Depends(get_current_user),  # ← Authentication required
):
    """
    Step 2.
    Receives the full list of questions with confirmed types from the frontend.
    Validates that every question has a type, then stores them in the session.
    Returns a summary ready for the frontend to display before generating.
    
    **Requires authentication**: Authorization header with Supabase JWT.
    **Validates ownership**: User must own the session they're trying to modify.
    """
    session_id = request.session_id
    
    if session_id not in _SESSIONS:
        raise HTTPException(
            status_code=404,
            detail="Session not found. Please upload your files again."
        )

    session = _SESSIONS[session_id]
    
    # Validate ownership — user must own this session
    if session.get("user_id") != user_id:
        raise HTTPException(
            status_code=403,
            detail="Access denied. You do not own this session."
        )

    questions = session["questions"]

    # Build confirmed_types dict from the submitted list
    confirmed_types = {item.q_text: item.q_type for item in request.question_types}

    # Validate — every question must resolve to a non-empty type
    missing = [
        q["q_text"][:80]
        for q in questions
        if not confirmed_types.get(q["q_text"], q.get("auto_type", ""))
    ]
    if missing:
        raise HTTPException(
            status_code=400,
            detail=(
                f"{len(missing)} question(s) still have no type assigned. "
                f"First few: {missing[:3]}"
            ),
        )

    # Persist in session
    session["confirmed_types"] = confirmed_types

    # Build summary
    type_counts: dict = {}
    for q in questions:
        t = confirmed_types.get(q["q_text"], q.get("auto_type") or "Single")
        type_counts[t] = type_counts.get(t, 0) + 1

    return {
        "status":       "confirmed",
        "session_id":   session_id,
        "total":        len(questions),
        "type_summary": type_counts,
        "message":      "Types confirmed. Ready to configure cuts and generate.",
    }


# ---------------------------------------------------------------------------
# ENDPOINT 3 — GENERATE
# ---------------------------------------------------------------------------

@app.post("/generate")
async def generate_databook(
    request: GenerateRequest,
    user_id: str = Depends(get_current_user),  # ← Authentication required
):
    """
    Step 3.
    Runs the full pipeline using the confirmed types and cuts from the request:
      1. _write_options_file()     — builds the clean options xlsx
      2. SurveyDatabookV2          — builds the base databook (unchanged class)
      3. apply_cuts_to_databook()  — applies demographic cuts (unchanged function)

    Returns the finished .xlsx as a file download.
    
    **Requires authentication**: Authorization header with Supabase JWT.
    **Validates ownership**: User must own the session they're trying to generate.
    """
    session_id = request.session_id
    
    if session_id not in _SESSIONS:
        raise HTTPException(
            status_code=404,
            detail="Session not found. Please upload your files again."
        )

    session = _SESSIONS[session_id]
    
    # Validate ownership — user must own this session
    if session.get("user_id") != user_id:
        raise HTTPException(
            status_code=403,
            detail="Access denied. You do not own this session."
        )

    questions       = session["questions"]
    confirmed_types = session["confirmed_types"]
    raw_data_path   = session["raw_data_path"]
    options_path    = session["options_path"]
    output_path     = session["output_path"]
    cuts_temp_path  = None

    try:
        # ── Write the clean options file ──────────────────────────────────────
        _write_options_file(questions, confirmed_types, options_path)

        # ── Build base databook (SurveyDatabookV2 — exactly as in merged script)
        processor = SurveyDatabookV2(raw_data_path, options_path)
        processor.load_options_file()
        processor.load_data()
        processor.identify_structure()
        processor.create_databook(output_path, survey_name=request.survey_name)

        # ── Apply cuts if any were provided ───────────────────────────────────
        if request.cuts:
            cuts_temp_path = _build_cuts_template(request.cuts)
            apply_cuts_to_databook(output_path, cuts_temp_path)

    except Exception as e:
        import traceback
        raise HTTPException(
            status_code=500,
            detail=f"Databook generation failed: {str(e)}\n\n{traceback.format_exc()}",
        )
    finally:
        # Always clean up the temporary cuts template
        if cuts_temp_path and os.path.exists(cuts_temp_path):
            os.unlink(cuts_temp_path)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename  = f"Databook_{timestamp}.xlsx"

    return FileResponse(
        path       = output_path,
        filename   = filename,
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers    = {"Content-Disposition": f"attachment; filename={filename}"},
    )


# ---------------------------------------------------------------------------
# ENDPOINT 4 — CLEANUP
# ---------------------------------------------------------------------------

@app.delete("/session/{session_id}")
async def cleanup_session(
    session_id: str,
    user_id: str = Depends(get_current_user),  # ← Authentication required
):
    """
    Call this after the user has downloaded their databook.
    Deletes all temp files for this session from disk.
    
    **Requires authentication**: Authorization header with Supabase JWT.
    **Validates ownership**: User must own the session they're trying to delete.
    """
    if session_id not in _SESSIONS:
        raise HTTPException(status_code=404, detail="Session not found.")
    
    session = _SESSIONS[session_id]
    
    # Validate ownership — user must own this session
    if session.get("user_id") != user_id:
        raise HTTPException(
            status_code=403,
            detail="Access denied. You do not own this session."
        )

    session_dir = os.path.dirname(_SESSIONS[session_id]["options_path"])
    shutil.rmtree(session_dir, ignore_errors=True)
    del _SESSIONS[session_id]

    return {"status": "cleaned", "session_id": session_id}


# ---------------------------------------------------------------------------
# STARTUP / SHUTDOWN
# ---------------------------------------------------------------------------

@app.on_event("startup")
async def _startup():
    os.makedirs(_TEMP_ROOT, exist_ok=True)
    print("=" * 60)
    print("  Survey Databook Generator API v2.0 — STARTED")
    print(f"  Temp dir : {_TEMP_ROOT}")
    print(f"  Docs     : http://localhost:8000/docs")
    print("=" * 60)


@app.on_event("shutdown")
async def _shutdown():
    shutil.rmtree(_TEMP_ROOT, ignore_errors=True)
    print("Survey Databook API — temp files cleaned up on shutdown.")
